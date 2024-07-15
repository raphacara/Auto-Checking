# ========================================= Import des bibliothèques ==============================================
import os
import sys
import json
import tempfile
import threading
import itertools

import tkinter as tk
from tkinter import filedialog, ttk, messagebox, font

import pandas as pd
import numpy as np
from PIL import Image, ImageTk
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList


# =============================================== CONFIGURATIONS ==================================================
bdd = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
file1 = file2 = rotated_image = preset = None

# Couleurs
color1 = '#012B65'
color2 = '#90CBFB'
color3 = '#FFC559'

# Polices
font_mini = ("Segoe UI", 9)
font_button = ("Segoe UI", 10, "bold")
font_style = ("Segoe UI", 15, "bold")
font_title = ("Segoe UI", 24, "bold")

# Entrées
ref_entry1 = date_entry1 = ref_entry2 = date_entry2 = ""
help_visible = False

# =============================================== Fonction Principale ===============================================
def merge_files():
    global ref_entry1, date_entry1, ref_entry2, date_entry2, checkbox_state1, checkbox_state2

    if not check_the_parameters():
        print("Wrong Parameters.")
        return
    print("Right Parameters !")

    name1 = os.path.basename(file1)
    name2 = os.path.basename(file2)
    expected_columns1 = [ref_entry1.lower().strip(), date_entry1.lower().strip()]
    expected_columns2 = [ref_entry2.lower().strip(), date_entry2.lower().strip()]

    progress_dialog = show_progress_dialog()

    try:
        df_file1 = read_excel_file(file1, expected_columns1)
    except ValueError as e:
        print(e)
        progress_dialog.destroy()
        messagebox.showwarning("Invalid Excel file",f"Columns not found in '{os.path.basename(file1)}'")
        return

    try:
        df_file2 = read_excel_file(file2, expected_columns2)
    except ValueError as e:
        print(e)
        progress_dialog.destroy()
        messagebox.showwarning("Invalid Excel file",f"Columns not found in'{os.path.basename(file2)}'")
        return

    cleanup(df_file1, df_file2)

    try:
        rename_columns(df_file1, ref_entry1, 'REFERENCE')
        rename_columns(df_file1, date_entry1, 'COMPARE')
        rename_columns(df_file2, ref_entry2, 'REFERENCE')
        rename_columns(df_file2, date_entry2, 'COMPARE')
    except ValueError as e:
        print(e)
        progress_dialog.destroy()
        return

    df_file1['REFERENCE'] = keep_eight_digits(df_file1['REFERENCE'])
    df_file2['REFERENCE'] = keep_eight_digits(df_file2['REFERENCE'])

    # Marquer les doublons dans les deux fichiers
    df_file1['is_duplicate'] = df_file1.duplicated(subset=['REFERENCE'], keep=False)
    df_file2['is_duplicate'] = df_file2.duplicated(subset=['REFERENCE'], keep=False)

    df_file1 = df_file1.drop_duplicates(subset=['REFERENCE'], keep='last')
    df_file2 = df_file2.drop_duplicates(subset=['REFERENCE'], keep='last')

    if df_file1['COMPARE'].dtype == 'datetime64[ns]':
        df_file1['COMPARE'] = pd.to_datetime(df_file1['COMPARE'], errors='coerce', format='%d/%m/%Y')
    if df_file2['COMPARE'].dtype == 'datetime64[ns]':
        df_file2['COMPARE'] = pd.to_datetime(df_file2['COMPARE'], errors='coerce', format='%d/%m/%Y')

    df_file1 = apply_date_conversion(df_file1, ['COMPARE'])
    df_file2 = apply_date_conversion(df_file2, ['COMPARE'])

    if checkbox_state1.get() and checkbox_state2.get():
        df_merged = pd.merge(df_file1, df_file2, on='REFERENCE', suffixes=('_1', '_2'), how='outer')
    elif checkbox_state1.get() and not checkbox_state2.get():
        df_merged = pd.merge(df_file1, df_file2, on='REFERENCE', suffixes=('_1', '_2'), how='left')
    elif checkbox_state2.get() and not checkbox_state1.get():
        df_merged = pd.merge(df_file1, df_file2, on='REFERENCE', suffixes=('_1', '_2'), how='right')
    else:
        df_merged = pd.merge(df_file1, df_file2, on='REFERENCE', suffixes=('_1', '_2'), how='inner')

    df_merged['Duplicates ?'] = (df_merged['is_duplicate_1'] | df_merged['is_duplicate_2'])
    df_merged['Duplicates ?'] = df_merged['Duplicates ?'].map({True: 'Yes', False: 'No'})

    # Logique pour marquer les lignes comme 'missing', 'identical' ou 'different'
    df_merged['Result'] = np.where(
        (df_merged['REFERENCE'].isna()) | (df_merged['REFERENCE'].isnull()),
        'missing',
        np.where(
            (df_merged['REFERENCE'].isin(df_file1['REFERENCE']) & ~df_merged['REFERENCE'].isin(df_file2['REFERENCE'])),
            f'missing in {name2}',
            np.where(
                (~df_merged['REFERENCE'].isin(df_file1['REFERENCE']) & df_merged['REFERENCE'].isin(df_file2['REFERENCE'])),
                f'missing in {name1}',
                np.where(
                    (df_merged['COMPARE_1'].isna() & df_merged['COMPARE_2'].isna()) | (df_merged['COMPARE_1'] == df_merged['COMPARE_2']),
                    'identical',
                    'different'
                )
            )
        )
    )

    # Insérer une croix rouge pour les cases de comparaison manquantes
    df_merged['COMPARE_1'] = np.where((df_merged['Result'] == f'missing in {name1}'), 'X', df_merged['COMPARE_1'])
    df_merged['COMPARE_2'] = np.where((df_merged['Result'] == f'missing in {name2}'), 'X', df_merged['COMPARE_2'])

    df_merged['COMPARE_1_datetime'] = pd.to_datetime(df_merged['COMPARE_1'], errors='coerce', format='%d.%m.%Y')
    df_merged['COMPARE_2_datetime'] = pd.to_datetime(df_merged['COMPARE_2'], errors='coerce', format='%d.%m.%Y')

    df_merged['Difference'] = (df_merged['COMPARE_2_datetime'] - df_merged['COMPARE_1_datetime']).dt.days

    df_output = df_merged[['REFERENCE', 'COMPARE_1', 'COMPARE_2', 'Result', 'Difference', 'Duplicates ?']].copy()

    df_output.rename(
        columns={'REFERENCE': ref_entry1_var.get(), 'COMPARE_1': f"{date_entry1_var.get()} \n({name1})",
                 'COMPARE_2': f"{date_entry2_var.get()} \n({name2})"}, inplace=True)

    progress_dialog.destroy()

    save_excel_with_customized_charts(df_output, selected_preset.get())

def check_the_parameters():
    if file1 is None or file2 is None:
        print("Please select both Excel files.")
        messagebox.showinfo("Error", "Please select both Excel files.")
        return False

    if selected_preset.get() == "Select your preset":
        print("Please select a preset")
        messagebox.showinfo("Error", "Please select a preset")
        return False

    if not all((ref_entry1, date_entry1, ref_entry2, date_entry2)):
        print("Please fill in all column names in the preset.")
        messagebox.showinfo("Error", "Please fill in all column names in the preset.")
        return False

    return True

def get_active_sheet_name(filepath):
    workbook = load_workbook(filename=filepath, read_only=True, keep_links=False)
    return workbook.active.title

def read_excel_file(filepath, expected_columns, sheet_name=None):
    if sheet_name is None:
        sheet_name = get_active_sheet_name(filepath)

    countmax = 0
    missing_columns = expected_columns  # Initialisation pour éviter l'erreur de référence avant assignation
    while countmax < 11:
        try:
            df = pd.read_excel(filepath, header=countmax, sheet_name=sheet_name)

            if df.empty:
                countmax += 1
                continue  # Si le DataFrame est vide, passer à l'itération suivante

            # Ignorer les lignes entièrement vides
            df.dropna(how='all', inplace=True)

            if df.empty:
                countmax += 1
                continue  # Si le DataFrame est vide après suppression des lignes vides, passer à l'itération suivante

            normalized_columns = [str(col).strip().lower() for col in df.columns]
            missing_columns = [col for col in expected_columns if col.lower().strip() not in normalized_columns]

            if not missing_columns:
                print(f"Found expected columns at header line {countmax}.")
                return df

            countmax += 1
        except Exception as e:
            raise ValueError(f"Error processing file '{filepath}': {str(e)}")

    if missing_columns:  # Vérifiez que missing_columns a été définie
        raise ValueError(
            f"Could not find the required columns in file '{filepath}'. Missing columns: {', '.join(missing_columns)}.")

def cleanup(df_file1, df_file2):
    global ref_entry1, date_entry1, ref_entry2, date_entry2
    # Nettoyage des noms de colonnes et des entrées de référence et de date
    ref_entry1 = ref_entry1.strip().lower()
    date_entry1 = date_entry1.strip().lower()
    ref_entry2 = ref_entry2.strip().lower()
    date_entry2 = date_entry2.strip().lower()
    df_file1.columns = [col.strip().lower() for col in df_file1.columns]
    df_file2.columns = [col.strip().lower() for col in df_file2.columns]
    print(df_file1.columns)

def rename_columns(df, original_name, new_name):
        if original_name in df.columns:
            df.rename(columns={original_name: new_name}, inplace=True)
        else:
            raise ValueError(f"Column {original_name} not found in the DataFrame")

def keep_eight_digits(series):
        series_as_str = series.astype(str)
        # Modifier pour extraire la première séquence de 8 chiffres ne commençant pas avec 0
        extracted = series_as_str.str.extract('([1-9]\d{7})')[0]
        # Remplace par la séquence extraite si elle est trouvée, sinon conserve la valeur originale
        return series.where(extracted.isna(), extracted)

def save_excel_with_customized_charts(df_output, sheet_name):
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if save_path:
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                temp_file_name = temp_file.name
                format_excel(df_output, temp_file_name, sheet_name)

            workbook = openpyxl.load_workbook(temp_file_name)
            sheet = workbook[sheet_name]

            # Ajouter les graphiques
            bar_chart = add_bar_chart(sheet)
            sheet.add_chart(bar_chart, "H2")

            # Vérifiez si la colonne "Difference" n'est pas vide avant d'ajouter le graphique en ligne
            if not df_output['Difference'].isnull().all():
                line_chart = add_line_chart(sheet)
                sheet.add_chart(line_chart, "H19")

            workbook.save(save_path)
            print("\nFormatted Excel file with customized charts saved successfully.")

        finally:
            if os.path.exists(temp_file_name):
                try:
                    os.unlink(temp_file_name)
                except PermissionError as e:
                    print(f"Erreur lors de la suppression du fichier temporaire: {e}")

        os.startfile(save_path)
        return save_path

    root.update()
    return None

def add_bar_chart(sheet):
    # Créer un graphique à barres pour les résultats 'identical', 'different', 'missing', 'duplicates'
    values = [0, 0, 0, 0]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
        if row[0].value == 'identical':
            values[0] += 1
        elif row[0].value == 'different':
            values[1] += 1
        elif row[0].value and row[0].value.startswith('missing'):
            values[2] += 1

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=6):
        if row[0].value == 'Yes':
            values[3] += 1

    data = [
        ['Category', 'Count'],
        ['Identical', values[0]],
        ['Different', values[1]],
        ['Missing', values[2]],
        ['Duplicates', values[3]]
    ]

    data_sheet = sheet.parent.create_sheet(title="TempSheetForBar")
    for row in data:
        data_sheet.append(row)

    chart = BarChart()
    chart.title = "Comparison Results"
    chart.style = 12
    chart.y_axis.title = 'Count'
    chart.x_axis.title = ''  # Enlever le titre de l'axe des catégories

    # Les couleurs et les catégories
    fill_colors = ["B0DBF1", "FFC7CE", "A9A9A9", "FF6666"]  # Identical, Different, Missing, Duplicates
    categories = ['Identical', 'Different', 'Missing', 'Duplicates']

    # Ajouter les données et appliquer les couleurs et les labels
    for inum, (color, category) in enumerate(zip(fill_colors, categories)):
        val_ref = Reference(data_sheet, min_row=1, max_row=len(categories) + 1, min_col=inum + 2, max_col=inum + 2)
        chart.add_data(val_ref, titles_from_data=True)  # Ajouter les données avec les titres

        series = chart.series[inum]
        series.graphicalProperties.solidFill = color

        # Ajouter le label à chaque série
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showCatName = False  # Désactiver les noms de catégories dans les étiquettes de données

    # Ajouter les catégories comme labels sous les barres
    cat_ref = Reference(data_sheet, min_col=1, min_row=2, max_row=len(categories) + 1)
    chart.set_categories(cat_ref)

    chart.legend = None  # Enlever la légende à droite

    # Ajuster les espacements et le chevauchement des barres pour les centrer
    chart.gapWidth = 150  # Ajuster l'espacement entre les barres (plus petite valeur pour moins d'espace)
    chart.overlap = 100  # Ajuster le chevauchement pour les centrer

    return chart

def add_line_chart(sheet):
    # Créer un graphique en ligne pour les différences
    difference_col = Reference(sheet, min_col=5, min_row=2, max_row=sheet.max_row)

    chart = LineChart()
    chart.style = 12
    chart.title = "Date differences overview"
    chart.y_axis.title = 'Number of days'
    # Enlever les labels de l'axe x
    chart.x_axis.delete = True

    chart.add_data(difference_col, titles_from_data=True)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = False

    # Ajuster les propriétés du graphique
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.y_axis.titleFont = Font(size=12, bold=True, color="000000")
    chart.legend = None  # Enlever la légende à droite

    return chart


# ======================================== Fonctions utilitaires =====================================================
def format_excel(df_output, excel_file, sheet_name):
    df_output.to_excel(excel_file, index=False)
    workbook = openpyxl.load_workbook(excel_file)
    ws = workbook.active
    ws.title = sheet_name  # Renommer la feuille active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    data_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    data_font = Font(color="000000")
    fill_identical = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_different = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_missing = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
    fill_red_bright = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    red_font = Font(color="FF0000", bold=True)

    max_width = 27
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = min(length, max_width)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = (adjusted_width + 2) * 1.2
        ws.row_dimensions[1].height = 43

        for cell in column_cells:
            cell.fill = data_fill
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    result_col_idx = None
    duplicates_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'Result':
            result_col_idx = idx
        if cell.value == 'Duplicates ?':
            duplicates_col_idx = idx

    if result_col_idx:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            result_cell = row[result_col_idx - 1]
            if result_cell.value == 'identical':
                fill = fill_identical
            elif result_cell.value == 'different':
                fill = fill_different
            else:
                fill = fill_missing
            for cell in row:
                cell.fill = fill
                if cell.value == 'X':
                    cell.font = red_font

    if duplicates_col_idx:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            duplicates_cell = row[duplicates_col_idx - 1]
            if duplicates_cell.value == 'Yes':
                duplicates_cell.fill = fill_red_bright  # Appliquer le remplissage rouge uniquement à la cellule "Yes"

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    workbook.save(excel_file)

def apply_date_conversion(df, date_columns):
    date_formats = ["%d/%m/%Y", "%d.%m.%Y", "%Y.%m.%d", "%Y/%m/%d"]

    def convert_dates(value):
        for date_format in date_formats:
            try:
                return pd.to_datetime(value, format=date_format, errors='raise').strftime('%d.%m.%Y')
            except (ValueError, TypeError):
                pass
        return value

    for col in date_columns:
        df[col] = df[col].apply(convert_dates)
    return df

def add_date_difference_column(df_merge):
    df_temp = df_merge.copy()
    df_merge['COMPARE_1'] = pd.to_datetime(df_merge['COMPARE_1'], format="%d/%m/%Y", errors='coerce')
    df_merge['COMPARE_2'] = pd.to_datetime(df_merge['COMPARE_2'], format="%d/%m/%Y", errors='coerce')
    df_temp['Date_difference'] = pd.NaT
    mask = df_merge['COMPARE_1'].notna() & df_merge['COMPARE_2'].notna()
    df_temp.loc[mask, 'Date_difference'] = (df_merge.loc[mask, 'COMPARE_2'] - df_merge.loc[mask, 'COMPARE_1']).dt.days
    return df_temp

def on_enter1(event):
    if file1 is None:
        event.widget.config(bg=color1, fg=color2)  # Changer la couleur lors du hover

def on_enter2(event):
    if file2 is None:
        event.widget.config(bg=color1, fg=color2)  # Changer la couleur lors du hover

def on_enter3(event):
    if file1 is None or file2 is None:
        event.widget.config(bg=color1, fg=color2)  # Changer la couleur lors du hover

def on_leave1(event):
    if file1 is None:
        event.widget.config(bg=color2, fg=color1)  # Revenir à la couleur d'origine

def on_leave2(event):
    if file2 is None:
        event.widget.config(bg=color2, fg=color1)  # Revenir à la couleur d'origine

def on_leave3(event):
    if file1 is None or file2 is None:
        event.widget.config(bg=color2, fg=color1)  # Revenir à la couleur d'origine

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f"{width}x{height}+{x}+{y}")

def update_image(image_label, image, angle, rotation, state):
    if not image_label.winfo_exists() or not state:
        return

    rotated_image2 = image.rotate(angle)  # Faire pivoter l'image
    rotated_image_tk = ImageTk.PhotoImage(rotated_image2)  # Convertir l'image en format Tkinter
    image_label.configure(image=rotated_image_tk)  # Mettre à jour l'image affichée dans le label
    image_label.image = rotated_image_tk  # Garder une référence à l'image pour éviter la suppression par le garbage collector
    root.after(20, update_image, image_label, image, angle-rotation, rotation, state)  # Appeler cette fonction après 200 millisecondes avec un angle augmenté

def generate_color_transition(start_hex, end_hex, steps):
    start_rgb = [int(start_hex[i1:i1+2], 16) for i1 in range(1, 6, 2)]
    end_rgb = [int(end_hex[i1:i1+2], 16) for i1 in range(1, 6, 2)]
    transition = [
        "#" + "".join(
            ["{:02x}".format(int(start_rgb[j] + ((end_rgb[j] - start_rgb[j]) * i1 / (steps-1))))
             for j in range(3)]
        ) for i1 in range(steps)
    ]
    return transition

def change_color():
    new_color = next(colors_cycle)
    bg_label.config(bg=new_color)
    root.after(200, change_color)  # Planifier le changement de couleur

def merge_files_with_progress_dialog():
    # Démarrer un thread pour exécuter la fonction merge_files en arrière-plan
    merge_thread = threading.Thread(target=merge_files)
    merge_thread.start()

def update_button_state():
    global ref_entry1, date_entry1, ref_entry2, date_entry2
    if file1:
        btn_select_file1.config(bg=color1, fg=color2)
    if file2:
        btn_select_file2.config(bg=color1, fg=color2)
    if file1 and file2 and selected_preset.get() != "Select your preset":
        btn_merge_files.config(state=tk.NORMAL, bg=color3,
                               fg=color1)  # Activer le bouton et changer la couleur de fond,
    else:
        btn_merge_files.config(state=tk.NORMAL, bg=color2,
                               fg="gray")  # Désactiver le bouton et changer la couleur de fond

    # Save the state
    save_last_state(file1, file2, selected_preset.get())

def save_button_data(button_index, title, ref_col_name_1, ref_col_name_2, date_col_name_1, date_col_name_2, checkbox_state_1, checkbox_state_2):
    button_datas = {
        "Title": title,
        "Ref_Column_Name_1": ref_col_name_1,
        "Ref_Column_Name_2": ref_col_name_2,
        "Date_Column_Name_1": date_col_name_1,
        "Date_Column_Name_2": date_col_name_2,
        "Checkbox_State_1": checkbox_state_1,
        "Checkbox_State_2": checkbox_state_2
    }
    # print(title, ref_col_name_1, date_col_name_1, ref_col_name_2, date_col_name_2, checkbox_state_1, checkbox_state_2)
    with open(f'button_{button_index}_data.txt', 'w') as fff:
        json.dump(button_datas, fff)

def update_selected_preset(the_preset, menu_color):
    global ref_entry1
    global ref_entry2
    global date_entry1
    global date_entry2
    global checkbox_state1
    global checkbox_state2

    print("Selected preset:", the_preset)

    # Vérifier si le preset est valide
    if the_preset not in button_texts:
        print(f"Preset {the_preset} is not valid.")
        return

    # Réinitialiser la couleur de tous les boutons
    for button in buttons_list:
        button.config(bg=color2, fg=color1)

    # Trouver l'index du preset sélectionné dans la liste des titres des boutons
    selected_index = button_texts.index(the_preset)

    # Changer la couleur du bouton associé au preset sélectionné
    buttons_list[selected_index].config(bg=color3, fg=color1)

    # Changer le texte sélectionné dans le menu déroulant
    selected_preset.set(button_texts[selected_index])

    # Changer la couleur du menu
    menu_color.config(bg=color1, font=font_style, fg=color2, bd=1)  # Appliquer les styles ici

    # Petite condition
    if file1 and file2 and selected_preset.get() != "Select your preset":
        btn_merge_files.config(bg=color3, fg=color1)  # Si les deux fichiers sont sélectionnés + le preset, le bouton devient jaune

    try:
        # Charger les données depuis le fichier texte correspondant au titre du preset sélectionné
        with open(f'button_{button_texts.index(the_preset) + 1}_data.txt', 'r') as f5:
            button_data_txt = json.load(f5)
            ref_entry1 = button_data_txt.get("Ref_Column_Name_1", "")
            date_entry1 = button_data_txt.get("Date_Column_Name_1", "")
            ref_entry2 = button_data_txt.get("Ref_Column_Name_2", "")
            date_entry2 = button_data_txt.get("Date_Column_Name_2", "")
            # Charger les états des cases à cocher
            checkbox_state1.set(button_data_txt.get("Checkbox_State_1", False))
            checkbox_state2.set(button_data_txt.get("Checkbox_State_2", False))
            # Mettre à jour les variables globales avec les nouvelles valeurs
            ref_entry1_var.set(ref_entry1)
            date_entry1_var.set(date_entry1)
            ref_entry2_var.set(ref_entry2)
            date_entry2_var.set(date_entry2)
            # Imprimer les valeurs mises à jour pour vérification
            print("ref1: ", ref_entry1, "\nref2: ", date_entry1, "\ndate1:", ref_entry2, "\ndate2:", date_entry2, "\nshow ref1: ",
                  checkbox_state1.get(), "\nshow ref2:", checkbox_state2.get())
    except FileNotFoundError:
        # Si le fichier du preset sélectionné n'est pas trouvé, laisser les variables inchangées
        print("Preset data file not found for", the_preset)

    save_last_state(file1, file2, the_preset)
    update_help_section()

    root.update()

def update_preset_menu():
    # Vide le menu déroulant actuel
    preset_menu['menu'].delete(0, 'end')

    # Liste pour stocker les nouveaux titres des boutons presets
    new_button_texts = []

    # Parcourez les index des boutons de 1 à 7 pour récupérer les titres des boutons presets
    for i7 in range(1, 8):
        try:
            # Ouvrez le fichier correspondant au bouton et chargez les données JSON
            with open(f'button_{i7}_data.txt', 'r') as f7:
                button_data7 = json.load(f7)
                title_button7 = button_data7.get("Title",
                                               f"Preset {i7}")  # Obtenez le titre du bouton ou utilisez "Preset {i}" par défaut
                new_button_texts.append(title_button7)  # Ajoutez le titre à la liste des textes de bouton
        except FileNotFoundError:
            # Si le fichier n'est pas trouvé, utilisez "Preset {i}" par défaut
            new_button_texts.append(f"Preset {i7}")

    # Mettre à jour le menu déroulant avec les nouveaux titres des boutons presets
    for txt2 in new_button_texts:
        preset_menu['menu'].add_command(label=txt2, command=lambda preset2=txt2: update_selected_preset(preset2, preset_menu))

    # Mettre à jour la liste button_texts
    button_texts[:] = new_button_texts

    # Si un preset était déjà sélectionné, le remettre à jour avec le nouveau texte
    selected_preset_text = selected_preset.get()
    if selected_preset_text in new_button_texts:
        selected_preset.set(selected_preset_text)
        save_last_state(file1, file2, selected_preset_text)
        print(selected_preset_text)
        update_help_section()

def save_last_state(file_1, file_2, preset_):
    if preset_ == "Select your preset":
        preset_ = ""
    state = {
        'file1': file_1,
        'file2': file_2,
        'preset': preset_
    }
    with open('last_state.txt', 'w') as f2:
        json.dump(state, f2)

def load_last_state():
    try:
        with open('last_state.txt', 'r') as f3:
            state = json.load(f3)
        return state
    except FileNotFoundError:
        return None

def reset():
    global file1, file2, preset
    file1 = file2 = preset = None
    selected_preset.set("Select your preset")
    preset_menu.config(bg=color2, font=font_style, fg=color1, bd=1)  # Appliquer les styles ici
    file_label1.grid_remove()  # Rendre le widget invisible au départ
    btn_select_file1.config(bg=color2, fg=color1)
    file_label2.grid_remove()  # Rendre le widget invisible au départ
    btn_select_file2.config(bg=color2, fg=color1)
    btn_select_file1.bind("<Enter>", lambda event: on_enter1(event))
    btn_select_file1.bind("<Leave>", lambda event: on_leave1(event))
    btn_select_file2.bind("<Enter>", lambda event: on_enter1(event))
    btn_select_file2.bind("<Leave>", lambda event: on_leave1(event))
    btn_merge_files.config(state=tk.DISABLED, bg=color2, fg="gray")

    update_button_state()
    update_help_section()
    save_last_state(file1, file2, selected_preset.get())

def help_step():
    global help_visible
    if not help_visible:
        help_frame.grid_remove()
    else:
        help_frame.grid(row=0, column=7, rowspan=7, sticky="nsew")
        update_help_section()
    help_visible = not help_visible

def open_video():
    video_path = os.path.join(bdd, 'example.mp4')
    if os.path.exists(video_path):
        os.startfile(video_path)  # Ouvrir le fichier avec l'application par défaut

# =========================================== Boutons de l'interface Graphique ======================================
def select_file1():
    global file1
    file1 = filedialog.askopenfilename(title="Select the 1st Excel file", filetypes=[("Excel Files", "*.xlsx; *.xlsb; *.xlsm; *.xls")])

    if file1:
        file_label1.config(text=os.path.basename(file1), bg=color2, fg=color1)
        update_button_state()  # Mettre à jour l'état du bouton
        file_label1.grid()  # Rendre le widget visible
    update_help_section()
    root.update()  # Mettre à jour l'interface utilisateur

def select_file2():
    global file2
    file2 = filedialog.askopenfilename(title="Select the 2nd Excel file", filetypes=[("Excel Files", "*.xlsx; *.xlsb; *.xlsm; *.xls")])
    if file2:
        file_label2.config(text=os.path.basename(file2), bg=color2, fg=color1)
        update_button_state()  # Mettre à jour l'état du bouton
        file_label2.grid()  # Rendre le widget visible
    update_help_section()
    root.update()  # Mettre à jour l'interface utilisateur

def show_progress_dialog():
    progress_dialog = tk.Toplevel()
    progress_dialog.title("Processing")
    progress_dialog.configure(background=color2)  # Définir la couleur d'arrière-plan
    progress_dialog.resizable(False, False)  # Désactiver la possibilité de redimensionner
    # progress_dialog.overrideredirect(True)  # Supprimer la barre de titre
    progress_dialog.attributes("-topmost", True)  # Garder la fenêtre au premier plan
    center_window(progress_dialog, 800, 400)  # Centrer la fenêtre sur l'écran
    progress_label = tk.Label(progress_dialog, text="Please wait while processing...", bg=color2, fg=color1, font=font_title)
    progress_label.pack(padx=20, pady=20)

    # Charger l'image
    path = os.path.join(bdd, 'engine.png')
    image = Image.open(path)
    image = image.resize((image.width // 3, image.height // 3))  # Redimensionner l'image
    image_tk = ImageTk.PhotoImage(image)

    # Créer un label pour afficher l'image
    image_label = tk.Label(progress_dialog, image=image_tk, bg=color2)
    image_label.pack()

    # Mettre à jour l'image affichée dans le label avec une rotation toutes les 500 millisecondes
    update_image(image_label, image, 90, 2, True)

    # Rendre la fenêtre modale
    # progress_dialog.transient(root)
    # progress_dialog.grab_set()
    # root.wait_window(progress_dialog)

    return progress_dialog

def open_preset_window(button_index):
    global selected_preset
    button = buttons_list[button_index - 1]  # -1, car les index commencent à 1, mais les listes commencent à 0.

    # Fonction pour effacer le contenu des zones de texte
    def clear_fields():
        for entry in entry_fields:
            entry.delete(0, tk.END)
        title_entry.delete(0, tk.END)  # Effacer le titre également
        checkbox_state1.set(False)
        checkbox_state2.set(False)

    # Fonction pour sauvegarder les données dans un fichier texte
    def save_data():
        # Récupérer les valeurs des entrées
        title = title_entry.get().strip()  # Obtenir le titre en supprimant les espaces inutiles
        if not title:  # Si le titre est vide
            title = f"Preset {button_index}"  # Utiliser le texte par défaut
        ref_col_name_1 = ref_col_entries[0].get()
        ref_col_name_2 = ref_col_entries[1].get()
        date_col_name_1 = date_col_entries[0].get()
        date_col_name_2 = date_col_entries[1].get()
        checkbox1_state = checkbox_state1.get()
        checkbox2_state = checkbox_state2.get()

        # Sauvegarder les données dans le fichier texte
        save_button_data(button_index, title, ref_col_name_1, ref_col_name_2, date_col_name_1, date_col_name_2,
                         checkbox1_state, checkbox2_state)

        # Mettre à jour le texte du bouton avec le titre entré par l'utilisateur
        button.config(text=title)

        # Mettre à jour le menu déroulant avec les nouveaux titres des boutons presets
        update_preset_menu()

        # Mettre à jour la sélection du menu déroulant
        selected_preset.set(button_texts[button_index - 1])

        # Fermer la fenêtre
        window.destroy()

        # Appeler la fonction pour mettre à jour le preset sélectionné
        update_selected_preset(title, preset_menu)

    # Créer une nouvelle fenêtre
    window = tk.Toplevel(root)
    window.configure(bg=color2)
    window.title("Button Details")
    window.geometry("300x264")
    window.update_idletasks()
    window.update_idletasks()  # Assurer que toutes les tâches en attente sont effectuées
    screen_w = window.winfo_screenwidth()
    screen_h = window.winfo_screenheight()
    window_w = window.winfo_width()
    window_h = window.winfo_height()
    x9 = (screen_w - window_w) // 2
    y9 = (screen_h - window_h) // 2
    window.geometry(f"+{x9}+{y9}")  # Positionner la fenêtre au centre de l'écran

    # Créer les widgets dans la fenêtre
    title_label = tk.Label(window, text="Preset title:", font=font_mini, bg=color2)
    title_label.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

    title_entry = tk.Entry(window, font=font_mini)
    title_entry.grid(row=0, column=1, columnspan=2, sticky="we", padx=10, pady=10)

    file_label = tk.Label(window, text="File 1                              File 2", font=font_button, bg=color2)
    file_label.grid(row=1, column=0, columnspan=3, sticky="nsew", padx=5, pady=10)

    ref_col_label = tk.Label(window, text="Reference column name:", font=font_mini, bg=color2)
    ref_col_label.grid(row=2, column=0, columnspan=3, sticky="nsew", padx=5, pady=0)

    date_col_label = tk.Label(window, text="Comparison column name:", font=font_mini, bg=color2)
    date_col_label.grid(row=5, column=0, columnspan=3, sticky="nsew", padx=5, pady=0)

    # Créer les widgets pour les colonnes de référence
    ref_col_entry1 = tk.Entry(window, font=font_mini)
    ref_col_entry1.grid(row=3, column=0, sticky="nsew", padx=10, pady=(2,0))

    ref_col_entry2 = tk.Entry(window, font=font_mini)
    ref_col_entry2.grid(row=3, column=1, sticky="nsew", padx=10, pady=(2,0))

    # Créer les cases à cocher pour la ligne 4
    checkbox1 = tk.Checkbutton(window, text="Show all references", variable=checkbox_state1, font=font_mini, bg=color2)
    checkbox1.grid(row=4, column=0, sticky="nsew", padx=10, pady=(0,10))

    checkbox2 = tk.Checkbutton(window, text="Show all references", variable=checkbox_state2, font=font_mini, bg=color2)
    checkbox2.grid(row=4, column=1, sticky="nsew", padx=10, pady=(0,10))

    # Créer les widgets pour les colonnes de date
    date_col_entry1 = tk.Entry(window, font=font_mini)
    date_col_entry1.grid(row=6, column=0, sticky="nsew", padx=10, pady=(2,10))

    date_col_entry2 = tk.Entry(window, font=font_mini)
    date_col_entry2.grid(row=6, column=1, sticky="nsew", padx=10, pady=(2,10))

    # Créer les listes contenant les widgets
    ref_col_entries = [ref_col_entry1, ref_col_entry2]
    date_col_entries = [date_col_entry1, date_col_entry2]

    # Créer une liste regroupant tous les champs d'entrée
    entry_fields = ref_col_entries + date_col_entries

    # Charger les données depuis le fichier texte s'il existe
    try:
        with open(f'button_{button_index}_data.txt', 'r') as f1:
            button_var = json.load(f1)
            title_entry.insert(0, button_var["Title"])
            ref_col_entries[0].insert(0, button_var["Ref_Column_Name_1"])
            ref_col_entries[1].insert(0, button_var["Ref_Column_Name_2"])
            date_col_entries[0].insert(0, button_var["Date_Column_Name_1"])
            date_col_entries[1].insert(0, button_var["Date_Column_Name_2"])
            checkbox_state1.set(button_var.get("Checkbox_State_1", False))
            checkbox_state2.set(button_var.get("Checkbox_State_2", False))
    except FileNotFoundError:
        pass

    # Créer les boutons "clear" et "save"
    clear_button = tk.Button(window, text="Clear", bg="red", fg="white", font=font_button, command=clear_fields)
    clear_button.grid(row=7, column=0, sticky="we", padx=10, pady=10)

    save_button = tk.Button(window, text="Save", bg="green", fg="white", font=font_button, command=save_data)
    save_button.grid(row=7, column=1, sticky="we", padx=10, pady=10)

    # Rendre la fenêtre modale
    window.transient(root)
    window.grab_set()
    root.wait_window(window)

def update_help_section():
    for widget in help_frame.winfo_children():
        widget.destroy()

    # Ajouter un titre au-dessus des étapes
    title_label = tk.Label(help_frame, text="Step-by-Step Guide", bg=color2, fg=color1, font=("Segoe UI", 12, "bold"))
    title_label.pack(fill='x', padx=10, pady=12)

    # Modifier la police pour inclure le soulignement
    underline_font = font.Font(title_label, title_label.cget("font"))
    underline_font.configure(underline=True)
    title_label.configure(font=underline_font)

    steps = [
        ("Select File 1", lambda: select_file1(), file1 is None),
        ("Select File 2", lambda: select_file2(), file2 is None),
        ("Select a Preset",
         lambda: preset_menu['menu'].post(580, 450),
         selected_preset.get() == "Select your preset" or selected_preset.get() == ""),
        ("Title for your comparison",
         lambda: open_preset_window(button_texts.index(selected_preset.get()) + 1),
         "Preset" in selected_preset.get()),
        ("Reference column in File 1",
         lambda: open_preset_window(button_texts.index(selected_preset.get()) + 1), ref_entry1 == ""),
        ("Reference column in File 2",
         lambda: open_preset_window(button_texts.index(selected_preset.get()) + 1), ref_entry2 == ""),
        ("Column to compare in File 1",
         lambda: open_preset_window(button_texts.index(selected_preset.get()) + 1), date_entry1 == ""),
        ("Column to compare in File 2",
         lambda: open_preset_window(button_texts.index(selected_preset.get()) + 1), date_entry2 == ""),
        ("Click on Fusion to compare!", lambda: btn_merge_files.invoke(),
         not (file1 is None or file2 is None or selected_preset.get() == "Select your preset" or any(
             not val for val in [ref_entry1, date_entry1, ref_entry2, date_entry2]))
         )
    ]

    previous_completed = True  # Initialement, on considère que la première étape est complétée

    for idx, (text, command, condition) in enumerate(steps):
        if previous_completed:
            if condition:
                buttons = tk.Button(help_frame, text=text, command=lambda cmd=command: [cmd(), update_help_section()], bg=color1, font=font_button, fg=color2, cursor="hand2")
                buttons.pack(fill='x', padx=10, pady=8)
                buttons.bind("<Enter>", lambda event, bouttonio=buttons: bouttonio.config(bg=color2, fg=color1))  # Ajouter hover
                buttons.bind("<Leave>", lambda event, bouttonio2=buttons: bouttonio2.config(bg=color1, fg=color2))  # Supprimer hover
                previous_completed = False  # Si une condition n'est pas remplie, les étapes suivantes ne seront pas affichées
            else:
                lbl = tk.Label(help_frame, text=text, bg=color2, fg=color1, font=font_button)
                lbl.pack(fill='x', padx=10, pady=8)
        else:
            break  # Si une étape n'est pas complétée, on arrête d'afficher les étapes suivantes

    # Ajouter l'auteur de l'application en bas
    footer_label = tk.Label(help_frame, text="Made by Raphaël CARABEUF", bg=color2, fg=color1, font=("Segoe UI", 6))
    footer_label.pack(side='bottom', fill='x', padx=10, pady=4)

    # Label pour le lien hypertexte
    link_label = tk.Label(help_frame, text="Example-Video", fg="blue", bg=color2, cursor="hand2", font=("Segoe UI", 13, "underline"))
    link_label.pack(side='bottom', fill='x', padx=0, pady=(8,16))
    link_label.bind("<Button-1>", lambda e: open_video())

# ================================================= Interface Graphique ============================================
root = tk.Tk()
root.title("Merge Excel Files")
root.geometry("1100x550")  # Augmentez la largeur pour inclure la section d'aide

# Définir l'icône de la fenêtre
icon_path = os.path.join(bdd, 'engine.ico')
root.iconbitmap(icon_path)

last_state = load_last_state()
if last_state:
    file1 = last_state.get('file1')
    file2 = last_state.get('file2')
    preset = last_state.get('preset')

# Créer un style personnalisé pour l'OptionMenu
style = ttk.Style(root)
style.theme_use('clam')  # Vous pouvez changer de thème si vous le souhaitez
style.configure("TMenubutton", background=color1, foreground=color2, font=font_style, padding=5)
style.map('TMenubutton', background=[('active', color3)], foreground=[('active', color1)])

ref_entry1_var = tk.StringVar()
date_entry1_var = tk.StringVar()
ref_entry2_var = tk.StringVar()
date_entry2_var = tk.StringVar()
checkbox_state1 = tk.BooleanVar()
checkbox_state2 = tk.BooleanVar()

# Configuration de la grille
for i in range(7):
    root.grid_rowconfigure(i, weight=1)  # Toutes les lignes ont le même poids
    root.grid_columnconfigure(i, weight=1)  # Toutes les colonnes ont le même poids
    root.grid_rowconfigure(i, weight=1)

root.grid_columnconfigure(0, weight=0, minsize=50)  # Ajustez 'minsize' pour définir la largeur
root.grid_columnconfigure(6, weight=0, minsize=50)  # Ajustez 'minsize' pour définir la largeur
root.grid_columnconfigure(3, weight=0, minsize=50)  # Ajustez 'minsize' pour définir la largeur

# Création des widgets
bg_path = os.path.join(bdd, 'intro1.png')
bg_image = Image.open(bg_path)
bg_img = ImageTk.PhotoImage(bg_image)  # Conversion en format compatible Tkinter

# Créer une étiquette pour l'image de fond
bg_label = tk.Label(root, image=bg_img, bg=color1)
bg_label.image = bg_img  # Empêcher le garbage collection de l'image
bg_label.grid(row=0, column=0, rowspan=7, columnspan=7, sticky="nsew")  # Spanning sur toute la grille
bg_label.lower()

# Démarrer l'animation de l'image de fond
# update_image(bg_label, bg_image, 90, 0.1, True)

# Ajouter une étiquette pour afficher le nom du fichier sélectionné
file_label1 = tk.Label(root, text="", font=font_mini, bg=color1, fg=color2, padx=0, pady=0, wraplength=100)
file_label1.grid(row=3, column=1, padx=40, pady=20, sticky="nsew")
file_label1.grid_remove()  # Rendre le widget invisible au départ

file_label2 = tk.Label(root, text="", font=font_mini, bg=color1, fg=color2, padx=0, pady=2, wraplength=100)
file_label2.grid(row=3, column=5, padx=40, pady=20, sticky="nsew")
file_label2.grid_remove()  # Rendre le widget invisible au départ

# Bouton Reset
btn_reset = tk.Button(root, text="Reset", command=reset, bg=color1, font=font_button,
                      relief=tk.SOLID, bd=1, fg=color2, padx=0, pady=0, cursor="target")
btn_reset.grid(row=0, column=0, ipadx=0, ipady=0, sticky="nsew")

# Bouton Do Nothing (nouveau bouton en haut à droite)
btn_help = tk.Button(root, text="Guide", command=help_step, bg=color1, font=font_button,
                     relief=tk.SOLID, bd=1, fg=color2, padx=0, pady=0, cursor="target")
btn_help.grid(row=0, column=6, ipadx=0, ipady=0, sticky="nsew")

# Ajouter une étiquette pour expliquer ce que fait l'application
app_explanation = tk.Label(root, text="Select two excel files to compare any column of your choice.",
                           font=font_style, bg=color1, fg=color2)
app_explanation.grid(row=0, column=1, columnspan=5, sticky="nsew")

# Bouton pour sélectionner le fichier 1
btn_select_file1 = tk.Button(root, text="Select 1st File", command=select_file1, bg=color2, font=font_style,
                             relief=tk.SOLID, bd=1, fg=color1, width=8, height=2, cursor="target", wraplength=100)
btn_select_file1.grid(row=2, column=1, padx=10, pady=(20, 0), sticky="nsew")
btn_select_file1.bind("<Enter>", lambda event: on_enter1(event))
btn_select_file1.bind("<Leave>", lambda event: on_leave1(event))

# Bouton pour sélectionner le fichier 2
btn_select_file2 = tk.Button(root, text="Select 2nd File", command=select_file2, bg=color2, font=font_style,
                             relief=tk.SOLID, bd=1, fg=color1, width=8, height=2, cursor="target", wraplength=100)
btn_select_file2.grid(row=2, column=5, padx=10, pady=(20, 0), sticky="nsew")
btn_select_file2.bind("<Enter>", lambda event: on_enter2(event))
btn_select_file2.bind("<Leave>", lambda event: on_leave2(event))

# Bouton pour lancer la fusion des données
btn_merge_files = tk.Button(root, text="Fusion", command=merge_files_with_progress_dialog, bg=color2, font=font_title,
                            relief=tk.SOLID, fg="gray", bd=1, width=8, height=1, cursor="target", wraplength=200)
btn_merge_files.grid(row=2, column=3, padx=100, pady=(20, 0), sticky="nsew")

# -------------------------------------------------- PRESETS -------------------------------------------------------------
toggle_button = tk.Button(root, text="▲  Presets  ▲", bg=color2, font=font_mini,
                          relief=tk.SOLID, bd=1, fg=color1, width=8, height=1, cursor="hand2")
toggle_button.grid(row=6, rowspan=2, column=0, columnspan=7, pady=(0, 0), sticky="nsew")
toggle_button.bind("<Enter>", lambda event: preset_frame.grid())

preset_frame = tk.Frame(root, bg=color1, bd=1, relief=tk.SOLID)
preset_frame.grid(row=6, column=0, columnspan=7, sticky="nsew")
preset_frame.grid_remove()  # Masquer par défaut

preset_label = tk.Label(preset_frame, text="Presets", font=font_style, bg=color1, fg=color2, bd=1)
preset_label.pack(fill="x")
preset_label.bind("<Leave>", lambda event: preset_frame.grid_remove())

# Créer les boutons dans chaque colonne avec un léger padding
button_texts = []  # Initialisez une liste vide pour stocker les titres des boutons

# Parcourez les index des boutons de 1 à 7.
for i in range(1, 8):
    try:
        # Ouvrez le fichier correspondant au bouton et chargez les données JSON
        with open(f'button_{i}_data.txt', 'r') as f:
            button_data = json.load(f)
            title_button = button_data.get("Title", f"Preset {i}")  # Obtenez le titre du bouton ou utilisez "Preset {i}" par défaut
            button_texts.append(title_button)  # Ajoutez le titre à la liste des textes de bouton
    except FileNotFoundError:
        # Si le fichier n'est pas trouvé, utilisez "Preset {i}" par défaut
        button_texts.append(f"Preset {i}")

# Création des boutons presets
buttons_list = []
for i, txt in enumerate(button_texts, start=1):
    btn = tk.Button(preset_frame, text=txt, bg=color2, font=font_mini, relief=tk.SOLID, bd=1, fg=color1, cursor="hand2", command=lambda index=i, text=txt: open_preset_window(index))
    btn.pack(side="left", fill="both", expand=True, padx=10, pady=10)
    buttons_list.append(btn)

# Menu déroulant
selected_preset = tk.StringVar(root)
selected_preset.set("Select your preset")  # Définir la valeur par défaut

# Créer le menu déroulant
preset_menu = tk.OptionMenu(root, selected_preset, *button_texts, command=lambda preset3: update_selected_preset(preset3, preset_menu))

# Utiliser le widget `OptionMenu` pour accéder au menu interne et appliquer les styles
menu = preset_menu.nametowidget(preset_menu.menuname)
menu.configure(font=font_style, bd=1, background=color1, foreground=color2, activebackground=color3, activeforeground=color1)

# Personnaliser le menu déroulant
preset_menu.config(bg=color2, font=font_style, fg=color1, bd=1)  # Appliquer les styles ici
preset_menu.grid(row=4, column=3, sticky="nsew", padx=70, pady=(100, 0))

# Ajouter un cadre pour la section d'aide
help_frame = tk.Frame(root, bg=color2, bd=2, relief=tk.SOLID, width=100)  # Définir la largeur fixe
help_frame.grid(row=0, column=7, rowspan=7, sticky="nsew")

# Initialiser la section d'aide comme cachée
help_visible = False
help_frame.grid_remove()

# Changer les couleurs du fond
colors = []
colors += generate_color_transition(color1, color2, 128)  # De color1 à color2
colors += generate_color_transition(color2, color3, 128)  # De color2 à color3
colors += generate_color_transition(color3, color1, 128)  # De color3 à color1
colors_cycle = itertools.cycle(colors)
root.after(500, change_color)  # Change la couleur du fond

# Appliquer l'état chargé après l'initialisation des widgets
if last_state:
    if file1:
        file_label1.config(text=os.path.basename(file1), bg=color2, fg=color1)
        update_button_state()
        file_label1.grid()

    if file2:
        file_label2.config(text=os.path.basename(file2), bg=color2, fg=color1)
        update_button_state()
        file_label2.grid()

    if preset:
        update_selected_preset(preset, preset_menu)

# Initialisation de la section d'aide
help_step()

# Fin
center_window(root, 1000, 500)  # Redimensionne la fenêtre
root.mainloop()  # démarrage de la boucle principale
